# !/Library/Frameworks/Python.framework/Versions/3.7/bin/python3
# -*- coding:utf-8 -*-
# @Author : Jiazhixiang
# @Time : 2019/12/11

'''
1.调用openpyxl.load_workbook()函数，打开“.xlsx”文件
2.获取“.xlsx”工作薄中名为“new title”的工作表。
3.sheetnames是用来获取工作薄所有工作表的名字的。如果你不知道工作薄到底有几个工作表，就可以把工作表的名字都打印出来。
4.把“new title”工作表中A1单元格赋值给A1_cell，再利用单元格value属性，就能打印出A1单元格的值。
'''

import openpyxl

wb = openpyxl.load_workbook('sample.xlsx')
ws = wb['new title']
sheetname = wb.sheetnames
print(sheetname)
A1_cell = ws['A1']
A1_value = A1_cell.value
print(A1_value)
