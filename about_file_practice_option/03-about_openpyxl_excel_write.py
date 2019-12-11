# !/Library/Frameworks/Python.framework/Versions/3.7/bin/python3
# -*- coding:utf-8 -*-
# @Author : Jiazhixiang
# @Time : 2019/12/11
import openpyxl
import datetime

wb = openpyxl.Workbook()  # 利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。
ws = wb.active
ws.title = 'new title'
ws['A1'] = 42
# ws.append([1, 2, 3])
# ws['A2'] = datetime.datetime.now()

# row = ['美国队长', '钢铁侠', '蜘蛛侠']
# ws.append(row)

rows = [['美国队长', '钢铁侠', '蜘蛛侠'], ['是', '漫威', '宇宙', '经典', '人物']]
for i in rows:
    ws.append(i)
wb.save("sample.xlsx")
